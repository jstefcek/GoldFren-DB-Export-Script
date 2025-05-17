import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter


def create_excel(export_data: dict, filename: str = "GoldFren_WebCatalog_V2.xlsx"):
    """
    Exports multiple datasets to an Excel file, with each dictionary key determining the sheet name.
    
    Params:
        export_data (dict): Dictionary containing multiple datasets (key = sheet name, value = list of objects).
        filename (str): Name of the output Excel file.
    """
    # Define output folder
    output_folder = r"C:\Users\Trix Gaming PC\Desktop\Goldfren\Scripts\GoldFren-DB-Export-Script\Excel Output"
    
    # Ensure the output directory exists
    os.makedirs(output_folder, exist_ok=True)
    
    file_path = os.path.join(output_folder, filename)

    # If file exists, delete it
    if os.path.exists(file_path): 
        os.remove(file_path)

    # Create Excel writer
    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        for sheet_name, data in export_data.items():
            # Convert object list to DataFrame
            data_list = [vars(item) for item in data.values()]
            df = pd.DataFrame(data_list)

            # Save DataFrame to a sheet with a dynamic name
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Load workbook to modify it
    wb = load_workbook(file_path)

    for sheet_name in export_data.keys():
        ws = wb[sheet_name]

        # Define Table range correctly even if columns > 26.
        num_rows, num_cols = ws.max_row, ws.max_column
        end_column = get_column_letter(num_cols)
        table_range = f"A1:{end_column}{num_rows}"

        table = Table(displayName=f"{sheet_name}_Table", ref=table_range)
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            # Add a little extra space for aesthetics
            ws.column_dimensions[col_letter].width = max_length + 2

    # Save workbook
    wb.save(file_path)
    print(f"[INFO] - Excel byl vytvoren do {file_path}")
